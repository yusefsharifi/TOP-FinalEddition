import pandas as pd
import xlsxwriter
from io import BytesIO
from typing import List, Dict, Any, Optional
from fastapi import Response
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import json
from datetime import datetime
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ExportManager:
    def __init__(self):
        self.supported_formats = ['pdf', 'excel', 'csv', 'json']
        
    def export_to_pdf(self, data: List[Dict[str, Any]], headers: List[str], title: str) -> BytesIO:
        """Export data to PDF format"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        # Add title
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30
        )
        elements.append(Paragraph(title, title_style))
        
        # Create table
        table_data = [headers] + [[str(row.get(header, '')) for header in headers] for row in data]
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        return buffer

    def export_to_excel(self, data: List[Dict[str, Any]], headers: List[str], title: str) -> BytesIO:
        """Export data to Excel format with styling"""
        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet title limit
        
        # Add title
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Add data
        for row_idx, row_data in enumerate(data, 4):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(row_data.get(header, '')))
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def export_to_csv(self, data: List[Dict[str, Any]], headers: List[str]) -> BytesIO:
        """Export data to CSV format"""
        buffer = BytesIO()
        writer = csv.DictWriter(buffer, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)
        buffer.seek(0)
        return buffer

    def export_to_json(self, data: List[Dict[str, Any]]) -> BytesIO:
        """Export data to JSON format"""
        buffer = BytesIO()
        json.dump(data, buffer, indent=2, ensure_ascii=False)
        buffer.seek(0)
        return buffer

    def get_export_response(self, data: List[Dict[str, Any]], headers: List[str], 
                          title: str, format: str) -> Response:
        """Get FastAPI Response object for export"""
        if format not in self.supported_formats:
            raise ValueError(f"Unsupported format: {format}")
        
        buffer = None
        content_type = None
        filename = f"{title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        if format == 'pdf':
            buffer = self.export_to_pdf(data, headers, title)
            content_type = 'application/pdf'
            filename += '.pdf'
        elif format == 'excel':
            buffer = self.export_to_excel(data, headers, title)
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename += '.xlsx'
        elif format == 'csv':
            buffer = self.export_to_csv(data, headers)
            content_type = 'text/csv'
            filename += '.csv'
        elif format == 'json':
            buffer = self.export_to_json(data)
            content_type = 'application/json'
            filename += '.json'
        
        return Response(
            content=buffer.getvalue(),
            media_type=content_type,
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"'
            }
        ) 