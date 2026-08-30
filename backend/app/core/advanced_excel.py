import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.drawing.image import Image
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional, Union, Tuple
from io import BytesIO
import qrcode
from datetime import datetime
import arabic_reshaper
from bidi.algorithm import get_display

class AdvancedExcelGenerator:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
    
    def set_sheet_name(self, name: str):
        """Set sheet name"""
        self.ws.title = name
    
    def add_title(self, title: str, font_size: int = 14, bold: bool = True):
        """Add title to the sheet"""
        # Reshape Persian text
        reshaped_text = arabic_reshaper.reshape(title)
        bidi_text = get_display(reshaped_text)
        
        cell = self.ws.cell(row=1, column=1, value=bidi_text)
        cell.font = Font(size=font_size, bold=bold)
        cell.alignment = Alignment(horizontal='center')
    
    def add_headers(self, headers: List[str], style: Optional[Dict[str, Any]] = None):
        """Add headers with styling"""
        default_style = {
            'font': Font(bold=True),
            'fill': PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid'),
            'alignment': Alignment(horizontal='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        
        style = style or default_style
        
        for col, header in enumerate(headers, 1):
            # Reshape Persian text
            reshaped_text = arabic_reshaper.reshape(header)
            bidi_text = get_display(reshaped_text)
            
            cell = self.ws.cell(row=3, column=col, value=bidi_text)
            for key, value in style.items():
                setattr(cell, key, value)
    
    def add_data(self, data: List[Dict[str, Any]], headers: List[str], 
                style: Optional[Dict[str, Any]] = None):
        """Add data with styling"""
        default_style = {
            'alignment': Alignment(horizontal='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        
        style = style or default_style
        
        for row_idx, row_data in enumerate(data, 4):
            for col_idx, header in enumerate(headers, 1):
                value = str(row_data.get(header, ''))
                # Reshape Persian text
                reshaped_text = arabic_reshaper.reshape(value)
                bidi_text = get_display(reshaped_text)
                
                cell = self.ws.cell(row=row_idx, column=col_idx, value=bidi_text)
                for key, value in style.items():
                    setattr(cell, key, value)
    
    def add_formula(self, formula: str, cell: str):
        """Add formula to a cell"""
        self.ws[cell] = formula
    
    def add_chart(self, chart_type: str, data_range: str, title: str, 
                  position: str = 'A1', size: Tuple[int, int] = (15, 10)):
        """Add chart to the sheet"""
        if chart_type == 'bar':
            chart = BarChart()
        elif chart_type == 'line':
            chart = LineChart()
        elif chart_type == 'pie':
            chart = PieChart()
        else:
            raise ValueError(f"Unsupported chart type: {chart_type}")
        
        chart.title = title
        chart.layout = Layout(
            manualLayout=ManualLayout(
                x=0.2, y=0.2,
                h=size[1], w=size[0]
            )
        )
        
        data = Reference(self.ws, range_string=data_range)
        chart.add_data(data)
        
        self.ws.add_chart(chart, position)
    
    def add_conditional_formatting(self, range_string: str, 
                                 rule_type: str, 
                                 formula: List[str],
                                 style: Dict[str, Any]):
        """Add conditional formatting"""
        from openpyxl.formatting.rule import FormulaRule
        
        rule = FormulaRule(
            formula=formula,
            stopIfTrue=True,
            **style
        )
        
        self.ws.conditional_formatting.add(range_string, rule)
    
    def add_data_validation(self, range_string: str, 
                          validation_type: str,
                          operator: str,
                          formula1: str,
                          formula2: Optional[str] = None):
        """Add data validation"""
        dv = DataValidation(
            type=validation_type,
            operator=operator,
            formula1=formula1,
            formula2=formula2,
            allow_blank=True
        )
        
        self.ws.add_data_validation(dv)
        dv.add(range_string)
    
    def add_qr_code(self, data: str, cell: str, size: Tuple[int, int] = (100, 100)):
        """Add QR code to the sheet"""
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Save QR code to buffer
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)
        
        # Add image to sheet
        excel_image = Image(buffer)
        excel_image.width = size[0]
        excel_image.height = size[1]
        self.ws.add_image(excel_image, cell)
    
    def protect_sheet(self, password: str):
        """Protect sheet with password"""
        self.ws.protection = SheetProtection(
            password=password,
            sheet=True,
            selectLockedCells=True,
            selectUnlockedCells=True,
            formatCells=True,
            formatColumns=True,
            formatRows=True,
            insertColumns=True,
            insertRows=True,
            insertHyperlinks=True,
            deleteColumns=True,
            deleteRows=True,
            sort=True,
            autoFilter=True,
            pivotTables=True
        )
    
    def auto_adjust_columns(self):
        """Auto-adjust column widths"""
        for column in self.ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            self.ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def generate_excel(self) -> BytesIO:
        """Generate Excel file"""
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer 