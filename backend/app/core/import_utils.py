import pandas as pd
from typing import List, Dict, Any, Optional
from fastapi import UploadFile
import json
import csv
from io import StringIO
import openpyxl
from datetime import datetime

class ImportManager:
    def __init__(self):
        self.supported_formats = ['excel', 'csv', 'json']
    
    async def import_from_excel(self, file: UploadFile) -> List[Dict[str, Any]]:
        """Import data from Excel file"""
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        
        # Get headers from first row
        headers = [str(cell.value).strip() for cell in ws[1]]
        
        # Get data
        data = []
        for row in ws.iter_rows(min_row=2):
            row_data = {}
            for header, cell in zip(headers, row):
                row_data[header] = cell.value
            data.append(row_data)
        
        return data
    
    async def import_from_csv(self, file: UploadFile) -> List[Dict[str, Any]]:
        """Import data from CSV file"""
        contents = await file.read()
        text = contents.decode('utf-8')
        csvfile = StringIO(text)
        reader = csv.DictReader(csvfile)
        return list(reader)
    
    async def import_from_json(self, file: UploadFile) -> List[Dict[str, Any]]:
        """Import data from JSON file"""
        contents = await file.read()
        return json.loads(contents)
    
    async def import_data(self, file: UploadFile, format: str) -> List[Dict[str, Any]]:
        """Import data from specified format"""
        if format not in self.supported_formats:
            raise ValueError(f"Unsupported format: {format}")
        
        if format == 'excel':
            return await self.import_from_excel(file)
        elif format == 'csv':
            return await self.import_from_csv(file)
        elif format == 'json':
            return await self.import_from_json(file)
    
    def validate_data(self, data: List[Dict[str, Any]], required_fields: List[str]) -> bool:
        """Validate imported data against required fields"""
        if not data:
            return False
        
        for row in data:
            for field in required_fields:
                if field not in row or row[field] is None:
                    return False
        
        return True
    
    def clean_data(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Clean imported data"""
        cleaned_data = []
        for row in data:
            cleaned_row = {}
            for key, value in row.items():
                if isinstance(value, str):
                    cleaned_row[key] = value.strip()
                elif isinstance(value, datetime):
                    cleaned_row[key] = value.isoformat()
                else:
                    cleaned_row[key] = value
            cleaned_data.append(cleaned_row)
        return cleaned_data 