from typing import List, Dict, Any, Optional, Union, Tuple
from io import BytesIO
import pandas as pd
import numpy as np
from datetime import datetime
import json
import csv
import xlsxwriter
from openpyxl import load_workbook
import arabic_reshaper
from bidi.algorithm import get_display
import logging
from pathlib import Path
import os
import hashlib
from cryptography.fernet import Fernet
from concurrent.futures import ThreadPoolExecutor
import re

class AdvancedImportExport:
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.encryption_key = Fernet.generate_key()
        self.cipher_suite = Fernet(self.encryption_key)
    
    def _validate_data(self, data: List[Dict[str, Any]], schema: Dict[str, Any]) -> Tuple[bool, List[str]]:
        """Validate data against schema"""
        errors = []
        for row_idx, row in enumerate(data, 1):
            for field, rules in schema.items():
                value = row.get(field)
                
                # Required field check
                if rules.get('required', False) and value is None:
                    errors.append(f"Row {row_idx}: {field} is required")
                    continue
                
                if value is not None:
                    # Type check
                    if 'type' in rules:
                        try:
                            if rules['type'] == 'int':
                                int(value)
                            elif rules['type'] == 'float':
                                float(value)
                            elif rules['type'] == 'date':
                                datetime.strptime(value, '%Y-%m-%d')
                            elif rules['type'] == 'email':
                                if not re.match(r"[^@]+@[^@]+\.[^@]+", value):
                                    raise ValueError("Invalid email format")
                        except (ValueError, TypeError) as e:
                            errors.append(f"Row {row_idx}: {field} has invalid type - {str(e)}")
                    
                    # Pattern check
                    if 'pattern' in rules:
                        if not re.match(rules['pattern'], str(value)):
                            errors.append(f"Row {row_idx}: {field} does not match pattern")
                    
                    # Range check
                    if 'range' in rules:
                        min_val, max_val = rules['range']
                        if not (min_val <= float(value) <= max_val):
                            errors.append(f"Row {row_idx}: {field} is out of range")
        
        return len(errors) == 0, errors
    
    def _clean_data(self, data: List[Dict[str, Any]], cleaning_rules: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Clean data according to rules"""
        cleaned_data = []
        for row in data:
            cleaned_row = {}
            for field, value in row.items():
                if field in cleaning_rules:
                    rules = cleaning_rules[field]
                    
                    # Trim whitespace
                    if rules.get('trim', False):
                        value = str(value).strip()
                    
                    # Convert case
                    if 'case' in rules:
                        if rules['case'] == 'upper':
                            value = str(value).upper()
                        elif rules['case'] == 'lower':
                            value = str(value).lower()
                    
                    # Remove special characters
                    if rules.get('remove_special_chars', False):
                        value = re.sub(r'[^a-zA-Z0-9\s]', '', str(value))
                    
                    # Replace values
                    if 'replace' in rules:
                        for old, new in rules['replace'].items():
                            value = str(value).replace(old, new)
                
                cleaned_row[field] = value
            cleaned_data.append(cleaned_row)
        return cleaned_data
    
    def _encrypt_data(self, data: str) -> str:
        """Encrypt data"""
        return self.cipher_suite.encrypt(data.encode()).decode()
    
    def _decrypt_data(self, encrypted_data: str) -> str:
        """Decrypt data"""
        return self.cipher_suite.decrypt(encrypted_data.encode()).decode()
    
    def export_to_excel(self, data: List[Dict[str, Any]], 
                       headers: List[str],
                       filename: str,
                       sheet_name: str = 'Sheet1',
                       style: Optional[Dict[str, Any]] = None) -> BytesIO:
        """Export data to Excel with advanced features"""
        buffer = BytesIO()
        workbook = xlsxwriter.Workbook(buffer)
        worksheet = workbook.add_worksheet(sheet_name)
        
        # Create styles
        header_style = workbook.add_format({
            'bold': True,
            'bg_color': '#CCCCCC',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        data_style = workbook.add_format({
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        # Write headers
        for col, header in enumerate(headers):
            # Reshape Persian text
            reshaped_text = arabic_reshaper.reshape(header)
            bidi_text = get_display(reshaped_text)
            worksheet.write(0, col, bidi_text, header_style)
        
        # Write data
        for row, row_data in enumerate(data, 1):
            for col, header in enumerate(headers):
                value = str(row_data.get(header, ''))
                # Reshape Persian text
                reshaped_text = arabic_reshaper.reshape(value)
                bidi_text = get_display(reshaped_text)
                worksheet.write(row, col, bidi_text, data_style)
        
        # Auto-adjust column widths
        for col in range(len(headers)):
            worksheet.set_column(col, col, 15)
        
        workbook.close()
        buffer.seek(0)
        return buffer
    
    def export_to_csv(self, data: List[Dict[str, Any]], 
                     headers: List[str],
                     filename: str) -> BytesIO:
        """Export data to CSV with advanced features"""
        buffer = BytesIO()
        writer = csv.DictWriter(buffer, fieldnames=headers)
        
        writer.writeheader()
        for row in data:
            # Reshape Persian text in values
            cleaned_row = {}
            for key, value in row.items():
                if key in headers:
                    reshaped_text = arabic_reshaper.reshape(str(value))
                    bidi_text = get_display(reshaped_text)
                    cleaned_row[key] = bidi_text
            writer.writerow(cleaned_row)
        
        buffer.seek(0)
        return buffer
    
    def export_to_json(self, data: List[Dict[str, Any]], 
                      filename: str,
                      indent: int = 2) -> BytesIO:
        """Export data to JSON with advanced features"""
        buffer = BytesIO()
        
        # Convert data to JSON with Persian text support
        json_data = []
        for row in data:
            cleaned_row = {}
            for key, value in row.items():
                # Reshape Persian text
                reshaped_text = arabic_reshaper.reshape(str(value))
                bidi_text = get_display(reshaped_text)
                cleaned_row[key] = bidi_text
            json_data.append(cleaned_row)
        
        json.dump(json_data, buffer, ensure_ascii=False, indent=indent)
        buffer.seek(0)
        return buffer
    
    def import_from_excel(self, file_path: str,
                         sheet_name: Optional[str] = None,
                         schema: Optional[Dict[str, Any]] = None,
                         cleaning_rules: Optional[Dict[str, Any]] = None) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Import data from Excel with validation and cleaning"""
        try:
            workbook = load_workbook(file_path, data_only=True)
            worksheet = workbook[sheet_name] if sheet_name else workbook.active
            
            # Get headers
            headers = []
            for cell in worksheet[1]:
                headers.append(str(cell.value))
            
            # Get data
            data = []
            for row in worksheet.iter_rows(min_row=2):
                row_data = {}
                for header, cell in zip(headers, row):
                    row_data[header] = str(cell.value) if cell.value is not None else None
                data.append(row_data)
            
            # Clean data if rules provided
            if cleaning_rules:
                data = self._clean_data(data, cleaning_rules)
            
            # Validate data if schema provided
            if schema:
                is_valid, errors = self._validate_data(data, schema)
                if not is_valid:
                    return [], errors
            
            return data, []
            
        except Exception as e:
            self.logger.error(f"Error importing Excel file: {str(e)}")
            return [], [str(e)]
    
    def import_from_csv(self, file_path: str,
                       schema: Optional[Dict[str, Any]] = None,
                       cleaning_rules: Optional[Dict[str, Any]] = None) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Import data from CSV with validation and cleaning"""
        try:
            data = []
            with open(file_path, 'r', encoding='utf-8') as file:
                reader = csv.DictReader(file)
                for row in reader:
                    data.append(row)
            
            # Clean data if rules provided
            if cleaning_rules:
                data = self._clean_data(data, cleaning_rules)
            
            # Validate data if schema provided
            if schema:
                is_valid, errors = self._validate_data(data, schema)
                if not is_valid:
                    return [], errors
            
            return data, []
            
        except Exception as e:
            self.logger.error(f"Error importing CSV file: {str(e)}")
            return [], [str(e)]
    
    def import_from_json(self, file_path: str,
                        schema: Optional[Dict[str, Any]] = None,
                        cleaning_rules: Optional[Dict[str, Any]] = None) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Import data from JSON with validation and cleaning"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
            
            # Clean data if rules provided
            if cleaning_rules:
                data = self._clean_data(data, cleaning_rules)
            
            # Validate data if schema provided
            if schema:
                is_valid, errors = self._validate_data(data, schema)
                if not is_valid:
                    return [], errors
            
            return data, []
            
        except Exception as e:
            self.logger.error(f"Error importing JSON file: {str(e)}")
            return [], [str(e)]
    
    def process_large_file(self, file_path: str,
                          chunk_size: int = 1000,
                          schema: Optional[Dict[str, Any]] = None,
                          cleaning_rules: Optional[Dict[str, Any]] = None) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Process large files in chunks"""
        try:
            all_data = []
            all_errors = []
            
            # Determine file type
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext == '.xlsx':
                workbook = load_workbook(file_path, data_only=True)
                worksheet = workbook.active
                
                # Get headers
                headers = []
                for cell in worksheet[1]:
                    headers.append(str(cell.value))
                
                # Process in chunks
                for row_idx in range(2, worksheet.max_row + 1, chunk_size):
                    chunk_data = []
                    for row in worksheet.iter_rows(min_row=row_idx, max_row=min(row_idx + chunk_size - 1, worksheet.max_row)):
                        row_data = {}
                        for header, cell in zip(headers, row):
                            row_data[header] = str(cell.value) if cell.value is not None else None
                        chunk_data.append(row_data)
                    
                    # Clean chunk if rules provided
                    if cleaning_rules:
                        chunk_data = self._clean_data(chunk_data, cleaning_rules)
                    
                    # Validate chunk if schema provided
                    if schema:
                        is_valid, errors = self._validate_data(chunk_data, schema)
                        if not is_valid:
                            all_errors.extend(errors)
                        else:
                            all_data.extend(chunk_data)
                    else:
                        all_data.extend(chunk_data)
            
            elif file_ext == '.csv':
                with open(file_path, 'r', encoding='utf-8') as file:
                    reader = csv.DictReader(file)
                    chunk_data = []
                    
                    for row in reader:
                        chunk_data.append(row)
                        if len(chunk_data) >= chunk_size:
                            # Clean chunk if rules provided
                            if cleaning_rules:
                                chunk_data = self._clean_data(chunk_data, cleaning_rules)
                            
                            # Validate chunk if schema provided
                            if schema:
                                is_valid, errors = self._validate_data(chunk_data, schema)
                                if not is_valid:
                                    all_errors.extend(errors)
                                else:
                                    all_data.extend(chunk_data)
                            else:
                                all_data.extend(chunk_data)
                            chunk_data = []
                    
                    # Process remaining data
                    if chunk_data:
                        if cleaning_rules:
                            chunk_data = self._clean_data(chunk_data, cleaning_rules)
                        
                        if schema:
                            is_valid, errors = self._validate_data(chunk_data, schema)
                            if not is_valid:
                                all_errors.extend(errors)
                            else:
                                all_data.extend(chunk_data)
                        else:
                            all_data.extend(chunk_data)
            
            return all_data, all_errors
            
        except Exception as e:
            self.logger.error(f"Error processing large file: {str(e)}")
            return [], [str(e)]
    
    def generate_import_log(self, file_path: str,
                          success_count: int,
                          error_count: int,
                          errors: List[str]) -> str:
        """Generate import operation log"""
        log = f"""
Import Operation Log
===================
File: {file_path}
Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Successfully imported records: {success_count}
Failed records: {error_count}

Errors:
{chr(10).join(f'- {error}' for error in errors)}
"""
        return log 